import pandas as pd
import openpyxl
import logging
from typing import List, Dict, Tuple, Optional
import re
from config import Config
from twilio_service import TwilioService

logger = logging.getLogger(__name__)

class ExcelHandler:
    def __init__(self):
        self.twilio_service = TwilioService()
        self.supported_extensions = ['.xlsx', '.xls']
    
    def read_excel_file(self, file_path: str) -> Tuple[pd.DataFrame, List[str]]:
        """Leer archivo Excel y devolver DataFrame con headers"""
        try:
            # Intentar leer el archivo
            df = pd.read_excel(file_path, header=None)
            
            # Obtener lista de columnas
            columns = [f"Columna {i+1}" for i in range(len(df.columns))]
            
            logger.info(f"Archivo Excel leído: {len(df)} filas, {len(df.columns)} columnas")
            return df, columns
            
        except Exception as e:
            logger.error(f"Error leyendo archivo Excel: {e}")
            raise Exception(f"Error leyendo archivo Excel: {str(e)}")
    
    def preview_data(self, df: pd.DataFrame, num_rows: int = 5) -> List[List]:
        """Obtener vista previa de los datos"""
        preview_data = []
        
        for idx, row in df.head(num_rows).iterrows():
            preview_data.append(row.tolist())
        
        return preview_data
    
    def extract_contacts(self, df: pd.DataFrame, phone_column_index: int,
                        column_mapping: Dict[str, int]) -> List[Dict]:
        """Extraer contactos del DataFrame"""
        contacts = []
        errors = []
        
        for idx, row in df.iterrows():
            try:
                # Obtener número de teléfono
                phone_raw = str(row.iloc[phone_column_index])
                
                # Validar y formatear número
                validation_result = self.twilio_service.validate_phone_number(phone_raw)
                
                if not validation_result['valid']:
                    errors.append(f"Fila {idx + 1}: {validation_result['error']} - {phone_raw}")
                    continue
                
                # Crear diccionario de contacto
                contact = {
                    'phone_number': validation_result['formatted']
                }
                
                # Agregar campos mapeados
                for field_name, col_index in column_mapping.items():
                    if col_index >= 0 and col_index < len(row):
                        value = row.iloc[col_index]
                        # Convertir NaN a None
                        if pd.isna(value):
                            contact[field_name] = None
                        else:
                            contact[field_name] = str(value).strip()
                
                # Agregar datos adicionales como JSON
                extra_data = {}
                for i, value in enumerate(row):
                    if i != phone_column_index and i not in column_mapping.values():
                        if not pd.isna(value):
                            extra_data[f"columna_{i+1}"] = str(value).strip()
                
                if extra_data:
                    contact['extra_data'] = pd.Series(extra_data).to_json()
                
                contacts.append(contact)
                
            except Exception as e:
                errors.append(f"Fila {idx + 1}: Error procesando - {str(e)}")
        
        logger.info(f"Contactos extraídos: {len(contacts)} válidos, {len(errors)} errores")
        
        return contacts, errors
    
    def standardize_phone_numbers(self, phone_numbers: List[str]) -> List[Dict]:
        """Estandarizar lista de números de teléfono"""
        results = []
        
        for phone in phone_numbers:
            validation = self.twilio_service.validate_phone_number(phone)
            results.append({
                'original': phone,
                'formatted': validation.get('formatted', ''),
                'valid': validation['valid'],
                'error': validation.get('error', '')
            })
        
        return results
    
    def get_column_statistics(self, df: pd.DataFrame) -> List[Dict]:
        """Obtener estadísticas de cada columna"""
        stats = []
        
        for col_idx in range(len(df.columns)):
            col_data = df.iloc[:, col_idx]
            
            # Contar valores no nulos
            non_null_count = col_data.count()
            
            # Detectar tipo de datos predominante
            data_types = col_data.dropna().apply(lambda x: type(x).__name__).value_counts()
            main_type = data_types.index[0] if len(data_types) > 0 else 'unknown'
            
            # Verificar si podría ser columna de teléfonos
            phone_pattern = re.compile(r'[\d\s\-\+\(\)]{8,}')
            potential_phones = col_data.astype(str).apply(
                lambda x: bool(phone_pattern.match(x)) if pd.notna(x) else False
            ).sum()
            
            stats.append({
                'column_index': col_idx,
                'column_name': f"Columna {col_idx + 1}",
                'non_null_count': int(non_null_count),
                'null_count': len(df) - non_null_count,
                'main_type': main_type,
                'unique_values': int(col_data.nunique()),
                'potential_phone_column': potential_phones > len(df) * 0.5,
                'sample_values': col_data.dropna().head(3).tolist()
            })
        
        return stats
    
    def validate_excel_file(self, file_path: str) -> Dict:
        """Validar archivo Excel antes de procesarlo"""
        try:
            # Verificar extensión
            if not any(file_path.lower().endswith(ext) for ext in self.supported_extensions):
                return {
                    'valid': False,
                    'error': 'Formato de archivo no soportado. Use .xlsx o .xls'
                }
            
            # Intentar abrir el archivo
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet = wb.active
            
            # Verificar si hay datos
            if sheet.max_row == 0 or sheet.max_column == 0:
                wb.close()
                return {
                    'valid': False,
                    'error': 'El archivo está vacío'
                }
            
            # Verificar tamaño
            if sheet.max_row > 10000:
                wb.close()
                return {
                    'valid': False,
                    'error': 'El archivo excede el límite de 10,000 filas'
                }
            
            wb.close()
            
            return {
                'valid': True,
                'rows': sheet.max_row,
                'columns': sheet.max_column
            }
            
        except Exception as e:
            return {
                'valid': False,
                'error': f'Error abriendo archivo: {str(e)}'
            }
    
    def export_contacts_to_excel(self, contacts: List[Dict], output_path: str):
        """Exportar contactos a archivo Excel"""
        try:
            # Convertir a DataFrame
            df = pd.DataFrame(contacts)
            
            # Reordenar columnas
            columns_order = ['phone_number', 'name', 'email', 'company']
            other_columns = [col for col in df.columns if col not in columns_order]
            final_order = columns_order + other_columns
            
            # Asegurarse de que todas las columnas existan
            for col in columns_order:
                if col not in df.columns:
                    df[col] = ''
            
            df = df[final_order]
            
            # Exportar a Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Contactos')
                
                # Ajustar ancho de columnas
                worksheet = writer.sheets['Contactos']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(f"Contactos exportados a: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exportando contactos: {e}")
            return False
